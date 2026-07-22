"""
Excel operation tools based on AWorld excel server.
Provides comprehensive Excel file manipulation capabilities.
"""
import json
import logging
from pathlib import Path
from typing import Dict, Any, List

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


# Export for other modules
__all__ = [
    'read_excel_data',
    'write_excel_data',
    'create_excel_workbook',
    'create_excel_worksheet',
    'apply_excel_formula',
    'get_excel_metadata',
    'create_excel_screenshot'
]


async def read_excel_data(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int = 1000
) -> Dict[str, Any]:
    """
    Read data from Excel file.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Specific sheet name (None for all sheets)
        max_rows: Maximum rows to read
        
    Returns:
        Dictionary with Excel data
    """
    try:
        path = Path(file_path).resolve()
        
        if not path.exists():
            return {"success": False, "error": f"File not found: {file_path}"}
        
        # Read Excel
        if sheet_name:
            df = pd.read_excel(path, sheet_name=sheet_name, nrows=max_rows)
            data = {sheet_name: df.to_dict(orient="records")}
            sheets = [sheet_name]
        else:
            excel_file = pd.ExcelFile(path)
            data = {}
            sheets = excel_file.sheet_names
            for sheet in sheets:
                df = pd.read_excel(path, sheet_name=sheet, nrows=max_rows)
                data[sheet] = df.to_dict(orient="records")
        
        return {
            "success": True,
            "file_path": str(path),
            "sheets": sheets,
            "data": data,
            "sheet_count": len(sheets)
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to read Excel: {str(e)}"}


async def write_excel_data(
    file_path: str,
    data: Dict[str, List[Dict]],
    overwrite: bool = False
) -> Dict[str, Any]:
    """
    Write data to Excel file.
    
    Args:
        file_path: Path to Excel file
        data: Dictionary of {sheet_name: [rows]}
        overwrite: Whether to overwrite existing file
        
    Returns:
        Dictionary with operation result
    """
    try:
        path = Path(file_path).resolve()
        
        if path.exists() and not overwrite:
            return {"success": False, "error": "File exists, use overwrite=True"}
        
        # Create Excel writer
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            for sheet_name, rows in data.items():
                df = pd.DataFrame(rows)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return {
            "success": True,
            "file_path": str(path),
            "sheets_written": len(data),
            "message": f"Wrote {len(data)} sheets to Excel"
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to write Excel: {str(e)}"}


async def create_excel_workbook(
    file_path: str
) -> Dict[str, Any]:
    """
    Create a new Excel workbook.
    
    Args:
        file_path: Path for new workbook
        
    Returns:
        Dictionary with result
    """
    try:
        path = Path(file_path).resolve()
        
        if path.exists():
            return {"success": False, "error": "File already exists"}
        
        wb = Workbook()
        wb.save(path)
        
        return {
            "success": True,
            "file_path": str(path),
            "message": "Created new workbook"
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to create workbook: {str(e)}"}


async def create_excel_worksheet(
    file_path: str,
    sheet_name: str
) -> Dict[str, Any]:
    """
    Create a new worksheet in Excel file.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name for new worksheet
        
    Returns:
        Dictionary with result
    """
    try:
        path = Path(file_path).resolve()
        
        if not path.exists():
            return {"success": False, "error": "File not found"}
        
        wb = load_workbook(path)
        
        if sheet_name in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' already exists"}
        
        wb.create_sheet(sheet_name)
        wb.save(path)
        
        return {
            "success": True,
            "file_path": str(path),
            "sheet_name": sheet_name,
            "message": f"Created worksheet '{sheet_name}'"
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to create worksheet: {str(e)}"}


async def apply_excel_formula(
    file_path: str,
    sheet_name: str,
    cell: str,
    formula: str
) -> Dict[str, Any]:
    """
    Apply formula to Excel cell.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Worksheet name
        cell: Cell reference (e.g., 'A1')
        formula: Excel formula (e.g., '=SUM(A1:A10)')
        
    Returns:
        Dictionary with result
    """
    try:
        path = Path(file_path).resolve()
        
        if not path.exists():
            return {"success": False, "error": "File not found"}
        
        wb = load_workbook(path)
        
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
        
        ws = wb[sheet_name]
        ws[cell] = formula
        wb.save(path)
        
        return {
            "success": True,
            "file_path": str(path),
            "sheet": sheet_name,
            "cell": cell,
            "formula": formula
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to apply formula: {str(e)}"}


async def get_excel_metadata(
    file_path: str
) -> Dict[str, Any]:
    """
    Get Excel file metadata.
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        Dictionary with metadata
    """
    try:
        path = Path(file_path).resolve()
        
        if not path.exists():
            return {"success": False, "error": "File not found"}
        
        wb = load_workbook(path, data_only=True)
        
        sheets_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets_info.append({
                "name": sheet_name,
                "max_row": ws.max_row,
                "max_column": ws.max_column
            })
        
        return {
            "success": True,
            "file_path": str(path),
            "file_size": path.stat().st_size,
            "sheets": sheets_info,
            "sheet_count": len(wb.sheetnames)
        }
        
    except Exception as e:
        return {"success": False, "error": f"Failed to get metadata: {str(e)}"}


async def create_excel_screenshot(
    file_path: str,
    sheet_name: str | None = None,
    output_dir: str = "."
) -> Dict[str, Any]:
    """
    Create a screenshot of Excel file (requires GUI environment).
    Note: This is a simplified implementation that exports to image.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Sheet to screenshot (None for first sheet)
        output_dir: Output directory for screenshot
        
    Returns:
        Dictionary with screenshot result
    """
    try:
        import time
        import sys
        import subprocess
        
        path = Path(file_path).resolve()
        
        if not path.exists():
            return {"success": False, "error": "File not found"}
        
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        timestamp = int(time.time())
        screenshot_file = output_path / f"{path.stem}_{sheet_name or 'sheet'}_{timestamp}.png"
        
        # This is a placeholder - actual implementation would require:
        # - pyautogui for screenshots
        # - or Excel API automation
        # - or conversion tools
        
        logger.info(f"📸 Excel screenshot would be saved to: {screenshot_file}")
        
        return {
            "success": False,
            "error": "Screenshot requires GUI environment and pyautogui",
            "note": "Install with: pip install pyautogui",
            "intended_output": str(screenshot_file)
        }
        
    except Exception as e:
        return {"success": False, "error": f"Screenshot failed: {str(e)}"}
